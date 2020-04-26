#! /usr/bin/env python3
import smtplib
import os
from multiprocessing.dummy import Pool
from datetime import datetime
from urllib.request import urlopen
from urllib.error import HTTPError
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import config_topnavigation as config


def topnavigation():
    ''' Get topnavigation links from vistaprint.com'''

    # set browser and url
    print('run browser...')
    driver = webdriver.Chrome()
    driver.get('http://vistaprint.com/')

    # find 'All Products' menu
    all_products = driver.find_element_by_xpath(
        '//*[@id="taxonomyItem14896"]/div/div/div')

    # get all menu and sub items
    menu_items = all_products.find_elements_by_tag_name('a')

    # get list with texts and hrefs from urls
    links = []
    print('get links')
    for link in menu_items:
        href = link.get_attribute('href')
        text = link.get_attribute('innerText')

        # Clear urls and texts
        text = text.strip()
        if '.aspx' in href:
            href = href.split('.aspx')
            href = href[0] + '.aspx'

        links.append([href, text])
    driver.close()
    return links


def save_result(links):
    ''' Save links to Excel '''

    print('save result')

    time = datetime.now()
    title = time.strftime('%m-%d-%Y')

    os.chdir(config.SAVE_FOLDER_TOPNAVIGATION)
    if os.path.exists(config.EXCEL_FILE_NAME):
        wb = load_workbook(config.EXCEL_FILE_NAME)
        ws = wb.create_sheet(title)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = title

    for row_nr, row in enumerate(links):
        for col_nr, col in enumerate(row):
            _ = ws.cell(column=col_nr+1, row=row_nr+1, value=col)
            last_row = row_nr+1

    last_cell = 'D' + str(last_row)
    ws[last_cell] = len(links)

    wb.save(config.EXCEL_FILE_NAME)


def get_status_code(link):
    ''' Return status code '''
    try:
        r = urlopen(link[0])
        status = str(r.getcode())
    except HTTPError as e:
        status = str(e)
    link.append(status)
    return link


def send_email(links):
    ''' Send alert email '''
    # Create list with 'bad' links
    bad_links = [link for link in links if link[2] != '200']

    # send email
    if bad_links:
        print('Send email')

        # Settings email
        email_from = config.EMAIL_FROM
        email_to = config.EMAIL_TO
        password = config.EMAIL_PASSWORD
        server = smtplib.SMTP(config.EMAIL_SERVER, config.EMAIL_PORT)

        # create message
        msg = ''
        for count, link in enumerate(bad_links):
            msg += '{}. {} -- {} -- {}\n'.format(count+1,
                                                 link[1],
                                                 link[2],
                                                 link[0])
        message = '\r\n'.join([
            'From:' + email_from,
            'To:' + email_to,
            'Subject: Alert - topnavigation status codes\n\n',
            msg
        ])

        # start email server and send message
        # server.starttls()
        server.login(email_from, password)
        server.sendmail(email_from, [email_to], message)
        server.quit

if __name__ == '__main__':
    links = topnavigation()

    print('Getting status codes')
    p = Pool(50)
    links_with_code = p.map(get_status_code, links)
    p.close()
    p.join()
    save_result(links_with_code)
    send_email(links_with_code)
