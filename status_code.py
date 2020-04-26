from datetime import datetime
from multiprocessing.dummy import Pool
import requests
from openpyxl import load_workbook
import config_status_code as config


def get_urls():
    ''' Return urls from Excel file '''
    print('Getting links')
    wb = load_workbook(config.EXCEL_FILE)
    ws = wb.worksheets[0]

    links = [cell.value for row in ws.rows for cell in row]
    return links


def get_codes(url):
    ''' Return status codes and list of redirects '''
    response = requests.get(url)
    row_for_excel = []
    if response.history:
        for rh in response.history:
            row_for_excel.append(rh.url)
            row_for_excel.append(rh.status_code)
    row_for_excel.append(response.url)
    row_for_excel.append(response.status_code)
    return row_for_excel


def save_data(rows):
    ''' Save data to Excel file '''
    print('Save file')
    wb = load_workbook(config.EXCEL_FILE)
    date = datetime.now()
    title = date.strftime('%m-%d-%Y')
    ws = wb.create_sheet(title)
    for row_nr, row in enumerate(rows):
        for col_nr, col in enumerate(row):
            _ = ws.cell(column=col_nr+1, row=row_nr+1, value=col)
    wb.save(config.EXCEL_FILE)


if __name__ == '__main__':
    t = datetime.now()
    all_urls = get_urls()
    all_urls = [url for url in all_urls if url is not None]

    print('Getting status codes')
    p = Pool(config.THREADS)
    data_for_excel = p.map(get_codes, all_urls)
    p.close()
    p.join
    save_data(data_for_excel)
    print(datetime.now() - t)
